import csv
import html
import json
import re
import time
from dataclasses import asdict, dataclass
from difflib import HtmlDiff, SequenceMatcher
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from PIL import Image, ImageChops, ImageDraw

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    Workbook = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except ImportError:
    SimpleDocTemplate = None

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.service import Service as ChromeService
    from selenium.webdriver.firefox.service import Service as FirefoxService
    from webdriver_manager.chrome import ChromeDriverManager
    from webdriver_manager.firefox import GeckoDriverManager
except ImportError:
    webdriver = None
    By = None
    ChromeService = None
    FirefoxService = None
    ChromeDriverManager = None
    GeckoDriverManager = None


REPORT_DIR = Path("reports")


@dataclass
class Defect:
    bug_id: str
    severity: str
    module: str
    description: str
    expected: str
    actual: str
    status: str = "Open"


@dataclass
class TestCaseResult:
    test_case: str
    result: str
    expected: str
    actual: str


@dataclass
class ApiCheck:
    method: str
    endpoint: str
    expected: str
    actual: str
    result: str
    response_time_seconds: float


class WebsiteComparer:
    SUPPORTED_BROWSERS = ("firefox", "chrome")

    def __init__(self, browser: str = "firefox"):
        REPORT_DIR.mkdir(exist_ok=True)
        self.browser = self.normalize_browser(browser)
        self.defects: List[Defect] = []
        self.test_cases: List[TestCaseResult] = []
        self.api_checks: List[ApiCheck] = []
        self.summary: Dict[str, object] = {}

    def normalize_browser(self, browser: str) -> str:
        browser = (browser or "firefox").strip().lower()
        if browser not in self.SUPPORTED_BROWSERS:
            supported = ", ".join(self.SUPPORTED_BROWSERS)
            raise ValueError(f"Unsupported browser '{browser}'. Choose one of: {supported}.")
        return browser

    def validate_url(self, url: str) -> bool:
        pattern = re.compile(
            r"^(https?://)"
            r"((([a-zA-Z0-9_-]+)\.)+[a-zA-Z]{2,}|localhost|127\.0\.0\.1)"
            r"(:[0-9]{1,5})?"
            r"(/.*)?$"
        )
        return bool(re.match(pattern, url.strip()))

    def normalize_url(self, url: str) -> str:
        url = url.strip()
        if not url.startswith(("http://", "https://")):
            return f"https://{url}"
        return url

    def check_accessibility(self, url: str) -> Tuple[bool, str]:
        try:
            response = requests.get(url, timeout=15)
            if response.status_code < 400:
                return True, f"HTTP {response.status_code}"
            return False, f"HTTP {response.status_code}"
        except requests.RequestException as exc:
            return False, str(exc)

    def create_driver(self):
        if webdriver is None:
            raise RuntimeError(
                "Selenium is not installed. Install dependencies with: "
                "pip install -r requirements.txt"
            )

        if self.browser == "chrome":
            options = webdriver.ChromeOptions()
            options.add_argument("--headless=new")
            options.add_argument("--disable-gpu")
            options.add_argument("--window-size=1366,768")
            service = ChromeService(ChromeDriverManager().install())
            return webdriver.Chrome(service=service, options=options)

        options = webdriver.FirefoxOptions()
        options.add_argument("--headless")
        service = FirefoxService(GeckoDriverManager().install())
        return webdriver.Firefox(service=service, options=options)

    def load_page(self, driver, url: str) -> float:
        started_at = time.perf_counter()
        driver.get(url)
        return round(time.perf_counter() - started_at, 2)

    def get_page_urls(self, url: str) -> List[str]:
        try:
            response = requests.get(url, timeout=15)
            response.raise_for_status()
        except requests.RequestException:
            return []

        soup = BeautifulSoup(response.text, "html.parser")
        page_urls: List[str] = []
        for link in soup.find_all("a"):
            href = link.get("href")
            if not href or href.startswith("#"):
                continue
            page_url = urljoin(url, href)
            if page_url not in page_urls:
                page_urls.append(page_url)
        return page_urls

    def check_broken_links(self, driver) -> List[Dict[str, str]]:
        broken_links: List[Dict[str, str]] = []
        links = driver.find_elements(By.TAG_NAME, "a")

        for link in links:
            href = link.get_attribute("href")
            if not href or href.startswith(("mailto:", "tel:", "javascript:")):
                continue

            try:
                response = requests.head(href, allow_redirects=True, timeout=10)
                if response.status_code >= 400:
                    broken_links.append({"url": href, "status": str(response.status_code)})
            except requests.RequestException as exc:
                broken_links.append({"url": href, "status": str(exc)})

        return broken_links

    def count_html_differences(self, html_a: str, html_b: str) -> int:
        matcher = SequenceMatcher(None, html_a.splitlines(), html_b.splitlines())
        return sum(1 for tag, _, _, _, _ in matcher.get_opcodes() if tag != "equal")

    def save_html_diff(self, html_a: str, html_b: str) -> int:
        diff = HtmlDiff()
        content_diff = diff.make_file(
            html_a.splitlines(),
            html_b.splitlines(),
            fromdesc="Website A",
            todesc="Website B",
        )
        Path("content_diff.html").write_text(content_diff, encoding="utf-8")
        (REPORT_DIR / "content_diff.html").write_text(content_diff, encoding="utf-8")
        return self.count_html_differences(html_a, html_b)

    def compare_screenshots(self, image_a_path: Path, image_b_path: Path) -> int:
        image_a = Image.open(image_a_path).convert("RGB")
        image_b = Image.open(image_b_path).convert("RGB")

        width = max(image_a.width, image_b.width)
        height = max(image_a.height, image_b.height)

        canvas_a = Image.new("RGB", (width, height), "white")
        canvas_b = Image.new("RGB", (width, height), "white")
        canvas_a.paste(image_a, (0, 0))
        canvas_b.paste(image_b, (0, 0))

        diff_image = ImageChops.difference(canvas_a, canvas_b)
        raw_diff_path = Path("visual_diff.png")
        diff_image.save(raw_diff_path)
        diff_image.save(REPORT_DIR / "visual_diff.png")

        bbox = diff_image.getbbox()
        if not bbox:
            return 0

        highlighted = canvas_b.copy()
        draw = ImageDraw.Draw(highlighted)
        draw.rectangle(bbox, outline="red", width=6)
        highlighted.save("visual_diff_highlighted.png")
        highlighted.save(REPORT_DIR / "visual_diff_highlighted.png")
        return 1

    def save_page_urls(self, url: str, page_urls: List[str], filename: str) -> None:
        path = Path(filename)
        lines = [f"Page URLs for {url}:", *page_urls]
        path.write_text("\n".join(lines), encoding="utf-8")
        (REPORT_DIR / filename).write_text("\n".join(lines), encoding="utf-8")

    def severity_for_broken_link(self, url: str) -> str:
        important_paths = ("checkout", "payment", "cart", "login", "signup", "order")
        if any(path in url.lower() for path in important_paths):
            return "High"
        return "Medium"

    def add_defect(
        self,
        severity: str,
        module: str,
        description: str,
        expected: str,
        actual: str,
    ) -> None:
        bug_id = f"BUG-{len(self.defects) + 1:03d}"
        self.defects.append(
            Defect(
                bug_id=bug_id,
                severity=severity,
                module=module,
                description=description,
                expected=expected,
                actual=actual,
            )
        )

    def add_test_case(self, name: str, passed: bool, expected: str, actual: str) -> None:
        self.test_cases.append(
            TestCaseResult(
                test_case=name,
                result="Pass" if passed else "Fail",
                expected=expected,
                actual=actual,
            )
        )

    def run_api_tests(self, base_url: Optional[str]) -> None:
        if not base_url:
            self.api_checks.append(
                ApiCheck("N/A", "API tests", "Optional API base URL", "Skipped", "Skipped", 0.0)
            )
            return

        base_url = base_url.rstrip("/")
        checks = [
            ("GET", "/products", None),
            ("GET", "/users", None),
            ("POST", "/login", {"email": "qa@example.com", "password": "Password123"}),
        ]

        for method, endpoint, payload in checks:
            started_at = time.perf_counter()
            try:
                if method == "GET":
                    response = requests.get(f"{base_url}{endpoint}", timeout=15)
                else:
                    response = requests.post(f"{base_url}{endpoint}", json=payload, timeout=15)

                elapsed = round(time.perf_counter() - started_at, 2)
                passed = response.status_code < 500
                self.api_checks.append(
                    ApiCheck(
                        method=method,
                        endpoint=endpoint,
                        expected="Status code below 500",
                        actual=f"HTTP {response.status_code}",
                        result="Pass" if passed else "Fail",
                        response_time_seconds=elapsed,
                    )
                )
                if not passed:
                    self.add_defect(
                        "High",
                        "API",
                        f"{method} {endpoint} returned server error.",
                        "API should return a non-5xx response.",
                        f"Received HTTP {response.status_code}.",
                    )
            except requests.RequestException as exc:
                elapsed = round(time.perf_counter() - started_at, 2)
                self.api_checks.append(
                    ApiCheck(method, endpoint, "Reachable API endpoint", str(exc), "Fail", elapsed)
                )
                self.add_defect(
                    "High",
                    "API",
                    f"{method} {endpoint} could not be tested.",
                    "API endpoint should be reachable.",
                    str(exc),
                )

    def compare_websites(self, url1: str, url2: str, api_base_url: Optional[str] = None) -> Dict[str, object]:
        url1 = self.normalize_url(url1)
        url2 = self.normalize_url(url2)

        valid_urls = self.validate_url(url1) and self.validate_url(url2) and url1 != url2
        self.add_test_case("URL Format Validation", valid_urls, "Two valid and different URLs", f"{url1}, {url2}")
        if not valid_urls:
            self.add_defect("Critical", "URL Validation", "Invalid or duplicate website URL provided.", "Two different valid URLs.", "Invalid or duplicate URL input.")
            self.export_reports()
            return self.summary

        accessible_a, access_message_a = self.check_accessibility(url1)
        accessible_b, access_message_b = self.check_accessibility(url2)
        self.add_test_case("Website A Accessible", accessible_a, "HTTP response below 400", access_message_a)
        self.add_test_case("Website B Accessible", accessible_b, "HTTP response below 400", access_message_b)

        if not accessible_a:
            self.add_defect("Critical", "Website Availability", "Website A is unreachable.", "Website A should be accessible.", access_message_a)
        if not accessible_b:
            self.add_defect("Critical", "Website Availability", "Website B is unreachable.", "Website B should be accessible.", access_message_b)
        if not accessible_a or not accessible_b:
            self.run_api_tests(api_base_url)
            self.export_reports()
            return self.summary

        driver_a = self.create_driver()
        driver_b = self.create_driver()

        try:
            load_time_a = self.load_page(driver_a, url1)
            load_time_b = self.load_page(driver_b, url2)

            broken_links_a = self.check_broken_links(driver_a)
            broken_links_b = self.check_broken_links(driver_b)

            html_a = driver_a.page_source
            html_b = driver_b.page_source
            html_difference_count = self.save_html_diff(html_a, html_b)

            screenshot_a = Path("screenshot1.png")
            screenshot_b = Path("screenshot2.png")
            driver_a.save_screenshot(str(screenshot_a))
            driver_b.save_screenshot(str(screenshot_b))
            (REPORT_DIR / "screenshot1.png").write_bytes(screenshot_a.read_bytes())
            (REPORT_DIR / "screenshot2.png").write_bytes(screenshot_b.read_bytes())

            visual_difference_count = self.compare_screenshots(screenshot_a, screenshot_b)

            page_urls_a = self.get_page_urls(url1)
            page_urls_b = self.get_page_urls(url2)
            self.save_page_urls(url1, page_urls_a, "website1_page_urls.txt")
            self.save_page_urls(url2, page_urls_b, "website2_page_urls.txt")

            total_links_checked = len(page_urls_a) + len(page_urls_b)
            total_broken_links = len(broken_links_a) + len(broken_links_b)

            self.add_test_case("HTML Match", html_difference_count == 0, "No HTML differences", f"{html_difference_count} difference group(s)")
            self.add_test_case("Broken Links", total_broken_links == 0, "No broken links", f"{total_broken_links} broken link(s)")
            self.add_test_case("Visual Match", visual_difference_count == 0, "No visual differences", f"{visual_difference_count} visual difference area(s)")
            self.add_test_case("Performance", load_time_a <= 3 and load_time_b <= 3, "Each site loads within 3 seconds", f"Site A: {load_time_a}s, Site B: {load_time_b}s")

            for broken_link in broken_links_a:
                self.add_defect(
                    self.severity_for_broken_link(broken_link["url"]),
                    "Website A Links",
                    f"Broken link found: {broken_link['url']}",
                    "All clickable links should return successful responses.",
                    f"Returned {broken_link['status']}.",
                )

            for broken_link in broken_links_b:
                self.add_defect(
                    self.severity_for_broken_link(broken_link["url"]),
                    "Website B Links",
                    f"Broken link found: {broken_link['url']}",
                    "All clickable links should return successful responses.",
                    f"Returned {broken_link['status']}.",
                )

            if html_difference_count:
                severity = "Low" if html_difference_count <= 3 else "Medium"
                self.add_defect(
                    severity,
                    "HTML Content",
                    f"HTML content differs between websites in {html_difference_count} group(s).",
                    "Both websites should have matching HTML structure/content for regression comparison.",
                    f"{html_difference_count} difference group(s) found. See content_diff.html.",
                )

            if visual_difference_count:
                self.add_defect(
                    "Medium",
                    "UI Visual Regression",
                    "Visual differences detected between website screenshots.",
                    "Website B should visually match Website A for the compared page.",
                    "Differences highlighted in visual_diff_highlighted.png.",
                )

            self.run_api_tests(api_base_url)

            self.summary = {
                "site_a": url1,
                "site_b": url2,
                "total_links_checked": total_links_checked,
                "broken_links": total_broken_links,
                "visual_differences": visual_difference_count,
                "html_differences": html_difference_count,
                "load_time_site_a": load_time_a,
                "load_time_site_b": load_time_b,
                "overall_status": "FAILED" if self.defects else "PASSED",
                "broken_links_a": broken_links_a,
                "broken_links_b": broken_links_b,
            }
        finally:
            driver_a.quit()
            driver_b.quit()

        self.export_reports()
        return self.summary

    def export_reports(self) -> None:
        if not self.summary:
            self.summary = {
                "site_a": "",
                "site_b": "",
                "total_links_checked": 0,
                "broken_links": 0,
                "visual_differences": 0,
                "html_differences": 0,
                "load_time_site_a": 0,
                "load_time_site_b": 0,
                "overall_status": "FAILED" if self.defects else "PASSED",
                "broken_links_a": [],
                "broken_links_b": [],
            }

        self.export_bug_reports()
        self.export_test_execution_report()
        self.export_api_report()
        self.export_dashboard()
        self.export_excel_report()
        self.export_uat_pdf()

    def write_csv(self, path: Path, rows: Iterable[Dict[str, object]], fieldnames: List[str]) -> None:
        with path.open("w", newline="", encoding="utf-8") as file:
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

    def export_bug_reports(self) -> None:
        fieldnames = ["bug_id", "severity", "module", "description", "expected", "actual", "status"]
        rows = [asdict(defect) for defect in self.defects]
        self.write_csv(Path("bug_report.csv"), rows, fieldnames)
        self.write_csv(REPORT_DIR / "bug_report.csv", rows, fieldnames)

        if Workbook is None:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Defects"
        ws.append(fieldnames)
        for row in rows:
            ws.append([row[field] for field in fieldnames])
        self.style_sheet(ws)
        wb.save("bug_report.xlsx")
        wb.save(REPORT_DIR / "bug_report.xlsx")

    def export_test_execution_report(self) -> None:
        fieldnames = ["test_case", "result", "expected", "actual"]
        rows = [asdict(test_case) for test_case in self.test_cases]
        self.write_csv(Path("test_execution_report.csv"), rows, fieldnames)
        self.write_csv(REPORT_DIR / "test_execution_report.csv", rows, fieldnames)

    def export_api_report(self) -> None:
        rows = [asdict(check) for check in self.api_checks]
        table_rows = "".join(
            "<tr>"
            f"<td>{html.escape(row['method'])}</td>"
            f"<td>{html.escape(row['endpoint'])}</td>"
            f"<td>{html.escape(row['expected'])}</td>"
            f"<td>{html.escape(row['actual'])}</td>"
            f"<td class='{row['result'].lower()}'>{html.escape(row['result'])}</td>"
            f"<td>{row['response_time_seconds']}</td>"
            "</tr>"
            for row in rows
        )
        content = self.html_page(
            "API Test Report",
            f"""
            <h1>API Test Report</h1>
            <table>
                <thead>
                    <tr><th>Method</th><th>Endpoint</th><th>Expected</th><th>Actual</th><th>Result</th><th>Time (s)</th></tr>
                </thead>
                <tbody>{table_rows}</tbody>
            </table>
            """,
        )
        Path("api_test_report.html").write_text(content, encoding="utf-8")
        (REPORT_DIR / "api_test_report.html").write_text(content, encoding="utf-8")

    def export_dashboard(self) -> None:
        defect_rows = "".join(
            "<tr>"
            f"<td>{html.escape(defect.bug_id)}</td>"
            f"<td><span class='severity {defect.severity.lower()}'>{html.escape(defect.severity)}</span></td>"
            f"<td>{html.escape(defect.module)}</td>"
            f"<td>{html.escape(defect.description)}</td>"
            f"<td>{html.escape(defect.status)}</td>"
            "</tr>"
            for defect in self.defects
        )
        test_rows = "".join(
            "<tr>"
            f"<td>{html.escape(test.test_case)}</td>"
            f"<td class='{test.result.lower()}'>{html.escape(test.result)}</td>"
            f"<td>{html.escape(test.expected)}</td>"
            f"<td>{html.escape(test.actual)}</td>"
            "</tr>"
            for test in self.test_cases
        )
        status_class = "failed" if self.summary["overall_status"] == "FAILED" else "passed"
        content = self.html_page(
            "QA Dashboard Report",
            f"""
            <h1>QA Dashboard Report</h1>
            <section class="summary-grid">
                {self.metric_card("Total Links Checked", self.summary["total_links_checked"])}
                {self.metric_card("Broken Links", self.summary["broken_links"])}
                {self.metric_card("Visual Differences", self.summary["visual_differences"])}
                {self.metric_card("HTML Differences", self.summary["html_differences"])}
                {self.metric_card("Site A Load Time", f'{self.summary["load_time_site_a"]} sec')}
                {self.metric_card("Site B Load Time", f'{self.summary["load_time_site_b"]} sec')}
            </section>
            <section class="overall {status_class}">
                <h2>Overall Status: {html.escape(str(self.summary["overall_status"]))}</h2>
            </section>
            <h2>Test Execution</h2>
            <table>
                <thead><tr><th>Test Case</th><th>Result</th><th>Expected</th><th>Actual</th></tr></thead>
                <tbody>{test_rows}</tbody>
            </table>
            <h2>Defects</h2>
            <table>
                <thead><tr><th>Bug ID</th><th>Severity</th><th>Module</th><th>Description</th><th>Status</th></tr></thead>
                <tbody>{defect_rows or '<tr><td colspan="5">No defects found.</td></tr>'}</tbody>
            </table>
            <h2>Artifacts</h2>
            <ul>
                <li><a href="content_diff.html">HTML Difference Report</a></li>
                <li><a href="visual_diff_highlighted.png">Highlighted Screenshot Difference</a></li>
                <li><a href="bug_report.csv">Bug Report CSV</a></li>
                <li><a href="qa_report.xlsx">QA Excel Report</a></li>
                <li><a href="api_test_report.html">API Test Report</a></li>
                <li><a href="UAT_Report.pdf">UAT Sign-off PDF</a></li>
            </ul>
            """,
        )
        Path("results.html").write_text(content, encoding="utf-8")
        (REPORT_DIR / "results.html").write_text(content, encoding="utf-8")

    def export_excel_report(self) -> None:
        if Workbook is None:
            return

        wb = Workbook()
        summary_ws = wb.active
        summary_ws.title = "Summary"
        summary_ws.append(["Metric", "Value"])
        for key, value in self.summary.items():
            if key not in ("broken_links_a", "broken_links_b"):
                summary_ws.append([key.replace("_", " ").title(), value])
        self.style_sheet(summary_ws)

        broken_ws = wb.create_sheet("Broken Links")
        broken_ws.append(["Website", "URL", "Status"])
        for row in self.summary.get("broken_links_a", []):
            broken_ws.append(["Website A", row["url"], row["status"]])
        for row in self.summary.get("broken_links_b", []):
            broken_ws.append(["Website B", row["url"], row["status"]])
        self.style_sheet(broken_ws)

        visual_ws = wb.create_sheet("Visual Differences")
        visual_ws.append(["Artifact", "Value"])
        visual_ws.append(["Visual Difference Count", self.summary["visual_differences"]])
        visual_ws.append(["Highlighted Screenshot", "visual_diff_highlighted.png"])
        self.style_sheet(visual_ws)

        performance_ws = wb.create_sheet("Performance")
        performance_ws.append(["Website", "Load Time Seconds"])
        performance_ws.append(["Website A", self.summary["load_time_site_a"]])
        performance_ws.append(["Website B", self.summary["load_time_site_b"]])
        self.style_sheet(performance_ws)

        defects_ws = wb.create_sheet("Defects")
        fields = ["bug_id", "severity", "module", "description", "expected", "actual", "status"]
        defects_ws.append(fields)
        for defect in self.defects:
            data = asdict(defect)
            defects_ws.append([data[field] for field in fields])
        self.style_sheet(defects_ws)

        wb.save("qa_report.xlsx")
        wb.save(REPORT_DIR / "qa_report.xlsx")

    def export_uat_pdf(self) -> None:
        if SimpleDocTemplate is None:
            return

        document = SimpleDocTemplate("UAT_Report.pdf", pagesize=A4)
        styles = getSampleStyleSheet()
        passed = sum(1 for test in self.test_cases if test.result == "Pass")
        failed = sum(1 for test in self.test_cases if test.result == "Fail")
        recommendation = "Approved for Release" if failed == 0 else "Not Approved - Defects must be resolved"

        elements = [
            Paragraph("UAT Sign-off Report", styles["Title"]),
            Spacer(1, 12),
            Paragraph(f"Application Name: Website Comparison Tool", styles["Normal"]),
            Paragraph(f"Version: Prototype QA Automation Upgrade", styles["Normal"]),
            Spacer(1, 12),
            Table(
                [
                    ["Test Cases Executed", len(self.test_cases)],
                    ["Passed", passed],
                    ["Failed", failed],
                    ["Open Defects", len(self.defects)],
                    ["Overall Status", self.summary["overall_status"]],
                    ["Recommendation", recommendation],
                ],
                colWidths=[180, 280],
            ),
            Spacer(1, 12),
            Paragraph("Defect Summary", styles["Heading2"]),
        ]

        defect_rows = [["Bug ID", "Severity", "Module", "Status"]]
        defect_rows.extend([[d.bug_id, d.severity, d.module, d.status] for d in self.defects])
        if len(defect_rows) == 1:
            defect_rows.append(["N/A", "N/A", "No defects found", "Closed"])

        table = Table(defect_rows, colWidths=[70, 80, 230, 80])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        elements.append(table)
        document.build(elements)
        (REPORT_DIR / "UAT_Report.pdf").write_bytes(Path("UAT_Report.pdf").read_bytes())

    def style_sheet(self, worksheet) -> None:
        header_fill = PatternFill("solid", fgColor="1F2937")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 60)

    def metric_card(self, label: str, value: object) -> str:
        return f"<article><span>{html.escape(label)}</span><strong>{html.escape(str(value))}</strong></article>"

    def html_page(self, title: str, body: str) -> str:
        return f"""<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{html.escape(title)}</title>
    <style>
        body {{ margin: 0; font-family: Arial, sans-serif; background: #f8fafc; color: #111827; }}
        main {{ max-width: 1120px; margin: 0 auto; padding: 32px 20px; }}
        h1 {{ margin: 0 0 24px; font-size: 32px; }}
        h2 {{ margin-top: 32px; font-size: 22px; }}
        table {{ width: 100%; border-collapse: collapse; background: #fff; margin-top: 12px; }}
        th, td {{ padding: 10px 12px; border: 1px solid #d1d5db; text-align: left; vertical-align: top; }}
        th {{ background: #1f2937; color: #fff; }}
        a {{ color: #0f766e; }}
        .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(190px, 1fr)); gap: 12px; }}
        .summary-grid article {{ background: #fff; border: 1px solid #d1d5db; border-radius: 8px; padding: 16px; }}
        .summary-grid span {{ display: block; color: #4b5563; font-size: 13px; }}
        .summary-grid strong {{ display: block; margin-top: 8px; font-size: 24px; }}
        .overall {{ margin-top: 20px; border-radius: 8px; padding: 16px; }}
        .overall.passed {{ background: #dcfce7; color: #166534; }}
        .overall.failed {{ background: #fee2e2; color: #991b1b; }}
        .pass {{ color: #166534; font-weight: bold; }}
        .fail {{ color: #991b1b; font-weight: bold; }}
        .skipped {{ color: #92400e; font-weight: bold; }}
        .severity {{ display: inline-block; padding: 3px 8px; border-radius: 999px; font-weight: bold; }}
        .critical {{ background: #7f1d1d; color: #fff; }}
        .high {{ background: #dc2626; color: #fff; }}
        .medium {{ background: #f59e0b; color: #111827; }}
        .low {{ background: #dbeafe; color: #1e3a8a; }}
    </style>
</head>
<body>
    <main>{body}</main>
</body>
</html>"""


def main() -> None:
    browser = input("Choose browser [firefox/chrome] (default firefox): ").strip() or "firefox"
    comparer = WebsiteComparer(browser=browser)
    url1 = input("Enter the first website URL: ")
    url2 = input("Enter the second website URL: ")
    api_base_url = input("Enter API base URL for optional API tests (press Enter to skip): ").strip()
    summary = comparer.compare_websites(url1, url2, api_base_url or None)
    print(json.dumps(summary, indent=2))
    print("Reports generated in the project root and reports/ folder.")


if __name__ == "__main__":
    main()
